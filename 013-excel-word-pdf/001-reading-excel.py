import openpyxl, os

os.chdir('C:\\Users\\Public\\code\\automate-with-python\\013-excel-word-pdf')

workbook = openpyxl.load_workbook('example.xlsx')

print(type(workbook))

sheet1 = workbook['Sheet1']

print(type(sheet1))

print(workbook.sheetnames)

cellA1 = sheet1['A1']

cell_value = cellA1.value

print(cellA1.value)


print(type(cell_value))

cellB1 = sheet1['B1']

cell_value = cellB1.value

print(cellB1.value)


print(type(cell_value))

cellC1 = sheet1['C1']

cell_value = cellC1.value

print(cellC1.value)


print(type(cell_value))

some_cell = sheet1.cell(row=1, column=2) # same thing as sheet1['B1']

print(some_cell.value)

##################

cellC8 = sheet1['C8']

cell_value = cellC8.value

print(cellC8.value)

print(type(cell_value))

numberOfRows = 0

for i in range(1, 100):
    some_cell = sheet1.cell(row=i, column=1)
    cell_value = some_cell.value
    if (cell_value is None):
        break
    numberOfRows += 1
    

print(numberOfRows)