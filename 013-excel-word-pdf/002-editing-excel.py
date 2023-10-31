import openpyxl, os

address = 'C:\\Users\\Public\\code\\automate-with-python\\013-excel-word-pdf'

wb = openpyxl.Workbook()

print(type(wb))

print(wb.sheetnames)

sheet1 = wb['Sheet']

print(type(sheet1))

print(sheet1['A1'].value == None)

sheet1['A1'] = 42
sheet1['A2'] = 'Hello'

os.chdir(address)

wb.save('edited_example.xlsx')

workbook = openpyxl.load_workbook('edited_example.xlsx')

print(type(workbook))

sheet1 = workbook['Sheet']

print(type(sheet1))

print(workbook.sheetnames)

cellA1 = sheet1['A1']

cell_value = cellA1.value

print(cellA1.value)


print(type(cell_value))

sheet2 = wb.create_sheet()

wb.save('edited_example.xlsx')

print(wb.sheetnames)

sheet2.title = 'Planilha 2'

wb.save('edited_example.xlsx')

wb.create_sheet(index=0, title='First Sheet 1')

wb.save('edited_example.xlsx')